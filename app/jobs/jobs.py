import copy
import os

from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
# from openpyxl.worksheet.table import Table, TableStyleInfo
from bot.db_commands import get_my_nc
from bot.loader import dp, bot


async def get_all():
    """Get all non_conformance with status 'см функцию get_my_nc' and send xlsx file"""
    group_chat = int(os.getenv("CHAT_ID"))
    all_nc = await get_my_nc()
    if all_nc:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Лист замечаний"
        non_conformance_list = []
        for nc in all_nc:
            id_nc = nc.pk
            created_at = nc.created_at
            formatted_datetime = created_at.strftime("%Y-%m-%d %H:%M:%S")
            description = nc.nc_description
            priority = nc.priority
            equipment = nc.equipment
            status = nc.status
            non_conformance_list.extend([id_nc, str(formatted_datetime), priority,
                                         str(equipment), description, str(status)])
        level = int(len(non_conformance_list) / 6)
        my_list = []
        x = 0
        while level != 0:
            lst = [item for item in non_conformance_list[x:(x + 6)]]
            my_list.append(lst)
            level -= 1
            x += 6

        ws1["A1"] = "ID"
        ws1["A1"].font = Font(name='Arial', bold=True, size=12, color="00FF0000")
        ws1["B1"] = "Дата"
        ws1["B1"].font = Font(name='Arial', bold=True, size=12, color="00FF0000")
        ws1.column_dimensions["B"].width = 18
        ws1["C1"] = "Группа"
        ws1["C1"].font = Font(name='Arial', bold=True, size=12, color="00FF0000")
        ws1.column_dimensions["C"].width = 10
        ws1["D1"] = "Оборудование"
        ws1["D1"].font = Font(name='Arial', bold=True, size=12, color="00FF0000")
        ws1.column_dimensions["D"].width = 25
        ws1["E1"] = "Описание"
        ws1["E1"].font = Font(name='Arial', bold=True, size=12, color="00FF0000")
        ws1.column_dimensions["E"].width = 35
        ws1['E1'].alignment = Alignment(wrap_text=True)
        ws1["F1"] = "Статус"
        ws1["F1"].font = Font(name='Arial', bold=True, size=12, color="00FF0000")
        ws1.column_dimensions["F"].width = 20
        # ws1.append(["ID", "Дата создания", "Группа", "Оборудование", "Описание", "Статус"])
        for i in range(6):
            for j in range(len(my_list)):
                _ = ws1.cell(column=i + 1, row=j + 2).value = my_list[j][i]

        # Делаем перенос в строках
        for row in ws1.iter_rows():
            for cell in row:
                alignment = copy.copy(cell.alignment)
                alignment.wrapText = True
                cell.alignment = alignment
        # height = len(my_list)
        ws1.auto_filter.ref = "A1:F1"
        wb.save("bot/breaking list.xlsx")

        return await bot.send_document(group_chat, open(r'bot/breaking list.xlsx', "rb"),
                                       caption='<b>Ежедневная рассылка несоответствий</b>')
    else:
        return await bot.send_message(chat_id=group_chat,
                                      text='<b>Нет доступных несоответствий</b>')
